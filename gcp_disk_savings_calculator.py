import os
import pandas as pd
from google.oauth2 import service_account
from googleapiclient import discovery
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Price reference table for US-West1 (Oregon)
# These prices are from Google Cloud pricing (May 2025)
DISK_PRICES = {
    'us-west1': {
        'pd-ssd': 0.17,      # per GB per month for SSD
        'pd-balanced': 0.10,  # per GB per month for Balanced
    },
    # Fallback for any other regions that might exist in your project
    'default': {
        'pd-ssd': 0.17,
        'pd-balanced': 0.10,
    }
}

def get_ssd_disk_data(project_id):
    """Collect all SSD persistent disk data from GCP project"""
    # Authenticate with Google Cloud
    # Replace with your preferred authentication method if needed
    compute = discovery.build('compute', 'v1')
    
    # Get all zones available in the project
    zones_response = compute.zones().list(project=project_id).execute()
    zones = [zone['name'] for zone in zones_response.get('items', [])]
    
    # List to store disk information
    disk_data = []
    
    # Filter for US-West1 zones only
    us_west1_zones = [zone for zone in zones if zone.startswith('us-west1-')]
    
    # For each zone in US-West1, get the disks
    for zone in us_west1_zones:
        try:
            # Get all disks in the zone
            disks_response = compute.disks().list(project=project_id, zone=zone).execute()
            disks = disks_response.get('items', [])
            
            for disk in disks:
                disk_name = disk['name']
                disk_type_url = disk['type']
                disk_type = disk_type_url.split('/')[-1]
                
                # Filter for SSD disks only
                if disk_type == 'pd-ssd':
                    disk_size_gb = int(disk['sizeGb'])
                    region = zone[:-2]  # Extract region from zone (e.g., us-central1 from us-central1-a)
                    
                    # Get VM attachment info (if attached)
                    attached_to = "Not attached"
                    is_boot = False
                    
                    if 'users' in disk:
                        # The disk is attached to one or more instances
                        for user in disk['users']:
                            # Extract instance name from URL
                            instance_name = user.split('/')[-1]
                            attached_to = instance_name
                            
                            # Check if it's a boot disk by examining the instance details
                            try:
                                instance = compute.instances().get(
                                    project=project_id,
                                    zone=zone,
                                    instance=instance_name
                                ).execute()
                                
                                for attached_disk in instance.get('disks', []):
                                    if attached_disk.get('source', '').endswith(disk_name) and attached_disk.get('boot', False):
                                        is_boot = True
                                        break
                            except Exception as e:
                                print(f"Error checking if disk is boot disk: {e}")
                            
                            break  # Just use the first attachment for simplicity
                    
                    # Add to our data list
                    disk_data.append({
                        'Disk Name': disk_name,
                        'Disk Type': disk_type,
                        'Disk Size (GB)': disk_size_gb,
                        'Zone': zone,
                        'Region': region,
                        'Attached To': attached_to,
                        'Is Boot Disk': 'Yes' if is_boot else 'No'
                    })
        
        except Exception as e:
            print(f"Error processing zone {zone}: {str(e)}")
    
    return disk_data

def calculate_savings(disk_data):
    """Calculate potential savings from converting SSD to Balanced disks"""
    savings_data = []
    
    for disk in disk_data:
        region = disk['Region']
        size_gb = disk['Disk Size (GB)']
        
        # Get pricing for this region, fallback to default if not found
        region_prices = DISK_PRICES.get(region, DISK_PRICES['default'])
        
        # Calculate current and potential costs
        current_cost = size_gb * region_prices['pd-ssd']
        balanced_cost = size_gb * region_prices['pd-balanced']
        monthly_savings = current_cost - balanced_cost
        annual_savings = monthly_savings * 12
        
        # Add savings info to the disk data
        savings_data.append({
            'Disk Name': disk['Disk Name'],
            'Disk Type': disk['Disk Type'],
            'Disk Size (GB)': size_gb,
            'Zone': disk['Zone'],
            'Region': region,
            'Attached To': disk['Attached To'],
            'Is Boot Disk': disk['Is Boot Disk'],
            'Current Monthly Cost (USD)': round(current_cost, 2),
            'Balanced Monthly Cost (USD)': round(balanced_cost, 2),
            'Monthly Savings (USD)': round(monthly_savings, 2),
            'Annual Savings (USD)': round(annual_savings, 2),
            'Savings Percentage': round((monthly_savings / current_cost) * 100, 1)
        })
    
    return savings_data

def create_excel_report(data, output_file="us_west1_disk_savings_report.xlsx"):
    """Create an Excel report with the disk savings data"""
    if not data:
        print("No SSD disks found.")
        return False
    
    # Convert to DataFrame
    df = pd.DataFrame(data)
    
    # Calculate totals
    total_monthly_savings = df['Monthly Savings (USD)'].sum()
    total_annual_savings = df['Annual Savings (USD)'].sum()
    total_current_cost = df['Current Monthly Cost (USD)'].sum()
    total_balanced_cost = df['Balanced Monthly Cost (USD)'].sum()
    total_disk_size = df['Disk Size (GB)'].sum()
    avg_savings_pct = (total_monthly_savings / total_current_cost) * 100 if total_current_cost > 0 else 0
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Disk Savings Report"
    
    # Add title
    ws.merge_cells('A1:L1')
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "US-West1 (Oregon) SSD to Balanced Disk Conversion Savings Report"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center')
    
    # Add summary section
    summary_row = 3
    ws.cell(row=summary_row, column=1).value = "Summary"
    ws.cell(row=summary_row, column=1).font = Font(bold=True)
    
    summary_data = [
        ("Total Disk Count:", len(data)),
        ("Total Disk Size (GB):", total_disk_size),
        ("Current Monthly Cost (USD):", round(total_current_cost, 2)),
        ("Projected Monthly Cost with Balanced Disks (USD):", round(total_balanced_cost, 2)),
        ("Potential Monthly Savings (USD):", round(total_monthly_savings, 2)),
        ("Potential Annual Savings (USD):", round(total_annual_savings, 2)),
        ("Average Savings Percentage:", f"{round(avg_savings_pct, 1)}%")
    ]
    
    for i, (label, value) in enumerate(summary_data):
        row = summary_row + i + 1
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=2).value = value
    
    # Add detailed data table
    detail_start_row = summary_row + len(summary_data) + 3
    
    # Write headers
    headers = [
        'Disk Name', 
        'Disk Size (GB)',
        'Is Boot Disk',
        'Attached To', 
        'Zone',
        'Region',
        'Current Monthly Cost (USD)',
        'Balanced Monthly Cost (USD)',
        'Monthly Savings (USD)',
        'Annual Savings (USD)',
        'Savings Percentage'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=detail_start_row, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="BFEFFF", end_color="BFEFFF", fill_type="solid")
    
    # Write data
    for row_num, disk in enumerate(data, detail_start_row + 1):
        ws.cell(row=row_num, column=1).value = disk['Disk Name']
        ws.cell(row=row_num, column=2).value = disk['Disk Size (GB)']
        ws.cell(row=row_num, column=3).value = disk['Is Boot Disk']
        ws.cell(row=row_num, column=4).value = disk['Attached To']
        ws.cell(row=row_num, column=5).value = disk['Zone']
        ws.cell(row=row_num, column=6).value = disk['Region']
        ws.cell(row=row_num, column=7).value = disk['Current Monthly Cost (USD)']
        ws.cell(row=row_num, column=8).value = disk['Balanced Monthly Cost (USD)']
        ws.cell(row=row_num, column=9).value = disk['Monthly Savings (USD)']
        ws.cell(row=row_num, column=10).value = disk['Annual Savings (USD)']
        ws.cell(row=row_num, column=11).value = f"{disk['Savings Percentage']}%"
    
    # Auto-adjust column widths
    for col_num, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = max(len(header) + 2, 15)
    
    # Save the workbook
    wb.save(output_file)
    print(f"Report saved to {output_file}")
    print(f"Total potential monthly savings: ${round(total_monthly_savings, 2)}")
    print(f"Total potential annual savings: ${round(total_annual_savings, 2)}")
    return True

def main():
    project_id = 'cas-prod-env'  # Use your project ID here
    
    print(f"Collecting SSD disk data for project: {project_id} in US-West1 (Oregon) region")
    disk_data = get_ssd_disk_data(project_id)
    
    if disk_data:
        print(f"Found {len(disk_data)} SSD disks in US-West1 region")
        savings_data = calculate_savings(disk_data)
        create_excel_report(savings_data)
    else:
        print("No SSD disks found in the project.")

if __name__ == "__main__":
    main()