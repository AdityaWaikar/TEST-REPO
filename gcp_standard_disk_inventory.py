import os
import pandas as pd
from google.oauth2 import service_account
from googleapiclient import discovery
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def get_vm_disk_data(project_id):
    """Collect VM and disk data from GCP project"""
    # Authenticate with Google Cloud
    # Replace with your own credentials method if needed
    # For example, using a service account key file:
    # credentials = service_account.Credentials.from_service_account_file('path/to/service_account.json')
    # Or use default credentials:
    # credentials, project = google.auth.default()
    
    # For this script, we'll use application default credentials
    # You need to run 'gcloud auth application-default login' beforehand
    compute = discovery.build('compute', 'v1')
    
    # Get all zones available in the project
    zones_response = compute.zones().list(project=project_id).execute()
    zones = [zone['name'] for zone in zones_response.get('items', [])]
    
    # List to store VM and disk information
    vm_disk_data = []
    
    # For each zone, get the instances
    for zone in zones:
        try:
            instances_response = compute.instances().list(project=project_id, zone=zone).execute()
            instances = instances_response.get('items', [])
            
            for instance in instances:
                vm_name = instance['name']
                
                # Process each disk attached to the VM
                for disk in instance.get('disks', []):
                    disk_name = disk['deviceName']
                    is_boot = disk['boot']
                    disk_type = None
                    disk_size_gb = None
                    
                    # Get the full disk information
                    disk_url = disk['source']
                    # Extract the disk name from the URL
                    disk_name_from_url = disk_url.split('/')[-1]
                    
                    # Get detailed disk information
                    disk_info = compute.disks().get(
                        project=project_id,
                        zone=zone,
                        disk=disk_name_from_url
                    ).execute()
                    
                    # Extract the disk type (last part of the URL path)
                    disk_type_url = disk_info['type']
                    disk_type = disk_type_url.split('/')[-1]
                    
                    # Check if it's Standard persistent disk
                    if disk_type == 'pd-standard':
                        disk_size_gb = int(disk_info['sizeGb'])
                        
                        # Add to our data list
                        vm_disk_data.append({
                            'VM Name': vm_name,
                            'Boot Disk': 'Yes' if is_boot else 'No',
                            'Disk Name': disk_name_from_url,
                            'Disk Type': disk_type,
                            'Disk Size (GB)': disk_size_gb,
                            'Zone': zone
                        })
        
        except Exception as e:
            print(f"Error processing zone {zone}: {str(e)}")
    
    return vm_disk_data

def create_excel_report(data, output_file="gcp_standard_disk_inventory.xlsx"):
    """Create an Excel report with the VM and disk data"""
    if not data:
        print("No VMs with standard persistent disks found.")
        return False
    
    # Convert to DataFrame for easier manipulation
    df = pd.DataFrame(data)
    
    # Create separate DataFrames for boot disks and additional disks
    boot_disks = df[df['Boot Disk'] == 'Yes'].copy()
    additional_disks = df[df['Boot Disk'] == 'No'].copy()
    
    # Create a Pandas Excel writer
    writer = pd.ExcelWriter(output_file, engine='openpyxl')
    
    # Get the workbook and create a new sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Standard Disk Inventory"
    
    # Write headers
    headers = [
        'VM Name', 
        'Boot Disk Name', 
        'Boot Disk Type', 
        'Boot Disk Size (GB)',
        'Additional Disk Name', 
        'Additional Disk Type', 
        'Additional Disk Size (GB)',
        'Zone'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="BFEFFF", end_color="BFEFFF", fill_type="solid")
    
    # Calculate column widths based on header length
    for col_num, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = max(len(header) + 2, 15)
    
    # Merge data from boot disks and additional disks
    vm_to_row = {}
    row_num = 2
    
    # Create a dictionary to track VMs and their assigned rows
    for _, boot_row in boot_disks.iterrows():
        vm_name = boot_row['VM Name']
        vm_to_row[vm_name] = row_num
        
        # Write VM info and boot disk info
        ws.cell(row=row_num, column=1).value = vm_name
        ws.cell(row=row_num, column=2).value = boot_row['Disk Name']
        ws.cell(row=row_num, column=3).value = boot_row['Disk Type']
        ws.cell(row=row_num, column=4).value = boot_row['Disk Size (GB)']
        ws.cell(row=row_num, column=8).value = boot_row['Zone']
        
        row_num += 1
    
    # Add additional disks info to existing VM rows or create new rows
    for _, add_row in additional_disks.iterrows():
        vm_name = add_row['VM Name']
        
        if vm_name in vm_to_row:
            # VM exists, add additional disk info
            row = vm_to_row[vm_name]
            
            # Check if this row already has an additional disk
            if ws.cell(row=row, column=5).value is not None:
                # Need a new row for this VM
                ws.cell(row=row_num, column=1).value = vm_name
                ws.cell(row=row_num, column=5).value = add_row['Disk Name']
                ws.cell(row=row_num, column=6).value = add_row['Disk Type']
                ws.cell(row=row_num, column=7).value = add_row['Disk Size (GB)']
                ws.cell(row=row_num, column=8).value = add_row['Zone']
                row_num += 1
            else:
                # Add to existing row
                ws.cell(row=row, column=5).value = add_row['Disk Name']
                ws.cell(row=row, column=6).value = add_row['Disk Type']
                ws.cell(row=row, column=7).value = add_row['Disk Size (GB)']
        else:
            # VM doesn't exist yet (additional disk without a boot disk)
            vm_to_row[vm_name] = row_num
            ws.cell(row=row_num, column=1).value = vm_name
            ws.cell(row=row_num, column=5).value = add_row['Disk Name']
            ws.cell(row=row_num, column=6).value = add_row['Disk Type']
            ws.cell(row=row_num, column=7).value = add_row['Disk Size (GB)']
            ws.cell(row=row_num, column=8).value = add_row['Zone']
            row_num += 1
    
    # Save the workbook
    wb.save(output_file)
    print(f"Report saved to {output_file}")
    return True

def main():
    project_id = 'cas-prod-env'  # Use your project ID here
    
    print(f"Collecting VM and standard persistent disk data for project: {project_id}")
    vm_disk_data = get_vm_disk_data(project_id)
    
    if vm_disk_data:
        print(f"Found {len(vm_disk_data)} instances of standard persistent disks")
        create_excel_report(vm_disk_data)
    else:
        print("No VMs with standard persistent disks found.")

if __name__ == "__main__":
    main()